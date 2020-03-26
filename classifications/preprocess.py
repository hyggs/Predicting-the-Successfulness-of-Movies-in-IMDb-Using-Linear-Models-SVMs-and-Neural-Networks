import numpy as nump
from openpyxl import load_workbook
from sklearn import preprocessing


def normalize(d):
    for i in range(len(d[0, :])):
        least = nump.min(d[:, i])   # 2
        largest = nump.max(d[:, i])     # 2781505847
        z = (d[:, i] - least) / (largest - least)
        d[:, i] = z
    for i in range(len(d[0, :])):
        least = nump.min(d[:, i])
        largest = nump.max(d[:, i])
        if i == 2:
            print(least)
            print(largest)
    return d


def pre_process():
    wb = load_workbook("tmdb_movies_data.xlsx")
    data = []
    ws = wb.worksheets[0]
    for row in ws.iter_rows():
        data.append([cell.value for cell in row])
    dataset = nump.array(data)
    # popularity, budget, revenue, runtime, vote_count, vote_average, budget_adj, revenue_adj, release_year
    columns = [2, 3, 4, 12, 16, 17, 19, 20, 18]
    d = dataset[1:, columns]
    mask = nump.ones(d.shape, dtype=bool)
    for i in range(len(d)):
        if 0 in d[i, :]:
            mask[i, :] = False
    d = d[mask]
    d = d.reshape(len(d) // len(columns), len(columns))
    # d now contains no rows that have 0 value.

    # d = normalize(d)

    return d    # 3855 movies


def pre_process_classed():
    wb = load_workbook("tmdb_movies_data.xlsx")
    data = []
    ws = wb.worksheets[0]
    for row in ws.iter_rows():
        data.append([cell.value for cell in row])
    dataset = nump.array(data)
    # popularity, budget, revenue, runtime, vote_count, vote_average, budget_adj, revenue_adj, release_year
    columns = [2, 3, 4, 12, 16, 17, 19, 20, 18]
    d = dataset[1:, columns]
    mask = nump.ones(d.shape, dtype=bool)
    for i in range(len(d)):
        if 0 in d[i, :]:
            mask[i, :] = False
    d = d[mask]
    d = d.reshape(len(d) // len(columns), len(columns))
    # d now contains no rows that have 0 value.

    # d = normalize(d)
    return d

